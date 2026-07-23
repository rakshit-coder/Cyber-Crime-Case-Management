"""
Generate comprehensive data dictionary with professional formatting.
Creates one Excel file with updated field specifications.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_professional_dictionary():
    """Create professional data dictionary in Excel format."""
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define colors and styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    overview_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    overview_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Overview Sheet
    overview = wb.create_sheet("Overview", 0)
    overview['A1'] = "Secure Case Connect - Database Schema"
    overview['A1'].font = Font(bold=True, size=14, color="1F4E78")
    overview['A3'] = "Table Name"
    overview['B3'] = "Field Count"
    overview['C3'] = "Description"
    
    for cell in ['A3', 'B3', 'C3']:
        overview[cell].fill = overview_fill
        overview[cell].font = overview_font
    
    tables_data = [
        ("users", 23, "User accounts and authentication"),
        ("activity_logs", 8, "User activity audit trail"),
        ("session_management", 8, "User session tracking"),
        ("complaints", 21, "Complaint registration and management"),
        ("evidence", 10, "Evidence related to complaints"),
        ("remarks", 10, "Remarks and comments on complaints"),
        ("complaint_attachments", 7, "File attachments for complaints"),
        ("case_workflow", 7, "Case processing workflow"),
        ("case_reports", 10, "Case investigation reports"),
        ("admin_notifications", 8, "Administrative notifications"),
        ("case_statistics", 8, "Case statistics and metrics"),
    ]
    
    for idx, (table_name, field_count, description) in enumerate(tables_data, start=4):
        overview[f'A{idx}'] = table_name
        overview[f'B{idx}'] = field_count
        overview[f'C{idx}'] = description
    
    overview.column_dimensions['A'].width = 25
    overview.column_dimensions['B'].width = 15
    overview.column_dimensions['C'].width = 40
    
    # Table Definitions with updated last_name
    tables_definition = {
        "users": [
            ("id", "INTEGER", "Primary Key", "Unique user identifier", "1, 2, 3"),
            ("username", "VARCHAR(30)", "Unique, Not Null", "Username for login", "john_doe, jane_smith"),
            ("email", "VARCHAR(30)", "Unique, Not Null", "User email address", "john@ex.com"),
            ("password", "VARCHAR(128)", "Not Null", "PBKDF2 hashed (CERT-IN: input 16-21 chars)", "[PBKDF2 hash]"),
            ("first_name", "VARCHAR(30)", "Null", "User first name", "John, Jane"),
            ("last_name", "VARCHAR(30)", "Null", "User last name", "Doe, Smith"),
            ("phone", "VARCHAR(15)", "Null", "Phone number", "+91-9876543210"),
            ("role", "VARCHAR(10)", "Choices: user, officer, admin", "User role in system", "user, officer, admin"),
            ("is_email_verified", "BOOLEAN", "Default: False", "Email verification status", "True, False"),
            ("email_verification_token", "VARCHAR(100)", "Null, Unique", "Email verification token", "[token]"),
            ("profile_picture", "VARCHAR(100)", "Null", "Profile picture file path", "[image_url]"),
            ("department", "VARCHAR(50)", "Null", "Officer department", "Cyber Crime, CERT-IN"),
            ("badge_number", "VARCHAR(30)", "Null", "Officer badge number", "CB-2026-001"),
            ("is_active", "BOOLEAN", "Default: True", "Account activation status", "True, False"),
            ("password_needs_reset", "BOOLEAN", "Default: False", "CERT-IN password reset flag", "True, False"),
            ("created_at", "DATETIME", "Auto-populated", "Account creation timestamp", "2026-02-24 10:30:00"),
            ("updated_at", "DATETIME", "Auto-updated", "Last modification timestamp", "2026-02-24 15:45:00"),
            ("is_staff", "BOOLEAN", "Default: False", "Django staff flag", "True, False"),
            ("is_superuser", "BOOLEAN", "Default: False", "Django superuser flag", "True, False"),
            ("last_login", "DATETIME", "Null", "Last login timestamp", "2026-02-24 14:20:00"),
            ("date_joined", "DATETIME", "Auto-populated", "Date joined", "2026-02-24 10:30:00"),
            ("groups", "ManyToMany", "Null", "User groups", "[group_ids]"),
        ],
        "activity_logs": [
            ("id", "INTEGER", "Primary Key", "Unique log identifier", "1, 2, 3"),
            ("user_id", "INTEGER", "Foreign Key (users)", "Reference to user", "1, 2, 3"),
            ("action", "VARCHAR(20)", "Choices: login, logout, etc", "Action performed", "login, logout, create_complaint"),
            ("description", "TEXT", "Not Null", "Detailed action description", "User logged in successfully"),
            ("ip_address", "VARCHAR(45)", "Null", "User IP address", "192.168.1.1, 10.0.0.5"),
            ("user_agent", "TEXT", "Null", "Browser/device info", "Mozilla/5.0 (Windows NT 10.0)"),
            ("action_hash", "VARCHAR(64)", "Not Null", "SHA-256 integrity hash", "[hash_value]"),
            ("created_at", "DATETIME", "Auto-populated, Indexed", "Log creation timestamp", "2026-02-24 10:30:00"),
        ],
        "session_management": [
            ("id", "INTEGER", "Primary Key", "Unique session identifier", "1, 2, 3"),
            ("user_id", "INTEGER", "Foreign Key (users)", "Reference to user", "1, 2, 3"),
            ("session_token", "VARCHAR(100)", "Unique, Not Null", "Session token", "[token_value]"),
            ("ip_address", "VARCHAR(45)", "Not Null", "Session IP address", "192.168.1.1"),
            ("user_agent", "TEXT", "Not Null", "Browser/device info", "Mozilla/5.0 (Windows NT 10.0)"),
            ("logged_in_at", "DATETIME", "Auto-populated", "Login timestamp", "2026-02-24 10:00:00"),
            ("last_activity", "DATETIME", "Auto-updated", "Last activity timestamp", "2026-02-24 15:00:00"),
            ("logged_out_at", "DATETIME", "Null", "Logout timestamp", "2026-02-24 16:00:00"),
            ("is_active", "BOOLEAN", "Default: True", "Session active status", "True, False"),
        ],
        "complaints": [
            ("id", "INTEGER", "Primary Key", "Unique complaint identifier", "1, 2, 3"),
            ("complaint_number", "VARCHAR(50)", "Unique, Not Null", "Complaint reference number", "COMP-2026-001"),
            ("complainant_name", "VARCHAR(30)", "Not Null", "Complainant full name", "John Doe"),
            ("complainant_email", "VARCHAR(30)", "Not Null", "Complainant email", "john@ex.com"),
            ("complainant_phone", "VARCHAR(15)", "Not Null", "Complainant phone number", "+91-9876543210"),
            ("description", "TEXT", "Not Null", "Complaint description", "Detailed complaint text"),
            ("incident_date", "DATE", "Not Null", "Incident date", "2026-02-20"),
            ("incident_location", "VARCHAR(30)", "Not Null", "Location of incident", "Delhi, Mumbai"),
            ("affected_system", "VARCHAR(30)", "Not Null", "System affected", "Email, Bank"),
            ("status", "VARCHAR(15)", "Choices: Submitted, In Review, etc", "Complaint status", "Submitted, In Review, Resolved"),
            ("approval_status", "VARCHAR(17)", "Choices: Pending, Approved, Rejected", "Approval status", "Pending, Approved, Rejected"),
            ("priority", "VARCHAR(10)", "Choices: Low, Medium, High, Critical", "Priority level", "High, Medium"),
            ("category", "VARCHAR(20)", "Choices: Cybercrime, Data Breach, etc", "Complaint category", "Cybercrime, Phishing"),
            ("assigned_officer_id", "INTEGER", "Foreign Key (users), Null", "Assigned officer", "2, 3"),
            ("analysis_report", "TEXT", "Null", "Preliminary analysis", "[analysis_text]"),
            ("created_at", "DATETIME", "Auto-populated", "Creation timestamp", "2026-02-24 10:00:00"),
            ("updated_at", "DATETIME", "Auto-updated", "Last update timestamp", "2026-02-24 15:00:00"),
            ("admin_notes", "TEXT", "Null", "Admin internal notes", "[notes_text]"),
            ("approval_date", "DATETIME", "Null", "Approval timestamp", "2026-02-25 10:00:00"),
            ("resolved_date", "DATETIME", "Null", "Resolution timestamp", "2026-02-28 16:00:00"),
            ("is_public", "BOOLEAN", "Default: False", "Public visibility flag", "True, False"),
        ],
        "evidence": [
            ("id", "INTEGER", "Primary Key", "Unique evidence identifier", "1, 2, 3"),
            ("complaint_id", "INTEGER", "Foreign Key (complaints)", "Reference to complaint", "1, 2, 3"),
            ("evidence_type", "VARCHAR(50)", "Not Null", "Type of evidence", "Screenshot, Document, Email"),
            ("description", "TEXT", "Not Null", "Evidence description", "Screenshot of fraudulent website"),
            ("file_path", "VARCHAR(255)", "Not Null", "File storage path", "[file_path]"),
            ("file_type", "VARCHAR(20)", "Choices: PDF, Image, Video, etc", "File MIME type", "PDF, JPG, PNG, MP4"),
            ("file_size", "INTEGER", "Not Null", "File size in bytes", "1024000"),
            ("uploaded_by_id", "INTEGER", "Foreign Key (users)", "Uploader reference", "1, 2"),
            ("created_at", "DATETIME", "Auto-populated", "Upload timestamp", "2026-02-24 11:00:00"),
            ("is_verified", "BOOLEAN", "Default: False, Indexed", "Evidence verification status", "True, False"),
        ],
        "remarks": [
            ("id", "INTEGER", "Primary Key", "Unique remark identifier", "1, 2, 3"),
            ("complaint_id", "INTEGER", "Foreign Key (complaints)", "Reference to complaint", "1, 2, 3"),
            ("officer_id", "INTEGER", "Foreign Key (users)", "Officer who made remark", "2, 3"),
            ("remark_text", "TEXT", "Not Null", "Remark content", "Further investigation required"),
            ("is_internal", "BOOLEAN", "Default: True", "Internal/External remark flag", "True, False"),
            ("is_public", "BOOLEAN", "Default: False", "Public visibility flag", "True, False"),
            ("priority_flag", "BOOLEAN", "Default: False", "Priority flag", "True, False"),
            ("created_at", "DATETIME", "Auto-populated", "Creation timestamp", "2026-02-24 12:00:00"),
            ("updated_at", "DATETIME", "Auto-updated", "Last update timestamp", "2026-02-24 14:00:00"),
            ("is_edited", "BOOLEAN", "Default: False", "Edit status flag", "True, False"),
        ],
        "complaint_attachments": [
            ("id", "INTEGER", "Primary Key", "Unique attachment identifier", "1, 2, 3"),
            ("complaint_id", "INTEGER", "Foreign Key (complaints)", "Reference to complaint", "1, 2, 3"),
            ("file_path", "VARCHAR(255)", "Not Null", "File storage path", "[file_path]"),
            ("attachment_type", "VARCHAR(20)", "Choices: Document, Image, etc", "File type", "Document, Image, Video"),
            ("file_size", "INTEGER", "Not Null", "File size in bytes", "512000"),
            ("uploaded_at", "DATETIME", "Auto-populated", "Upload timestamp", "2026-02-24 10:30:00"),
            ("is_verified", "BOOLEAN", "Default: False", "Verification status", "True, False"),
        ],
        "case_workflow": [
            ("id", "INTEGER", "Primary Key", "Unique workflow identifier", "1, 2, 3"),
            ("complaint_id", "INTEGER", "Foreign Key (complaints)", "Reference to complaint", "1, 2, 3"),
            ("current_stage", "VARCHAR(50)", "Not Null", "Current workflow stage", "Investigation, Escalation"),
            ("action", "VARCHAR(25)", "Choices: Assign, Review, Escalate", "Action taken", "Assign, Review, Escalate"),
            ("performed_by_id", "INTEGER", "Foreign Key (users)", "Officer who performed action", "2, 3"),
            ("performed_at", "DATETIME", "Auto-populated", "Action timestamp", "2026-02-24 13:00:00"),
            ("notes", "TEXT", "Null", "Action notes", "Case assigned for investigation"),
        ],
        "case_reports": [
            ("id", "INTEGER", "Primary Key", "Unique report identifier", "1, 2, 3"),
            ("complaint_id", "INTEGER", "Foreign Key (complaints)", "Reference to complaint", "1, 2, 3"),
            ("report_title", "VARCHAR(150)", "Not Null", "Report title", "Cyber Fraud Investigation Report"),
            ("report_content", "TEXT", "Not Null", "Detailed report content", "[report_text]"),
            ("findings", "TEXT", "Not Null", "Investigation findings", "[findings_text]"),
            ("recommendations", "TEXT", "Null", "Recommended actions", "[recommendations]"),
            ("created_by_id", "INTEGER", "Foreign Key (users)", "Report author", "2, 3"),
            ("created_at", "DATETIME", "Auto-populated", "Report creation timestamp", "2026-02-25 10:00:00"),
            ("updated_at", "DATETIME", "Auto-updated", "Last update timestamp", "2026-02-26 15:00:00"),
            ("is_final", "BOOLEAN", "Default: False", "Report finalization status", "True, False"),
        ],
        "admin_notifications": [
            ("id", "INTEGER", "Primary Key", "Unique notification identifier", "1, 2, 3"),
            ("admin_id", "INTEGER", "Foreign Key (users)", "Target admin user", "3, 4"),
            ("notification_type", "VARCHAR(50)", "Not Null", "Type of notification", "New Complaint, Case Update"),
            ("message", "TEXT", "Not Null", "Notification message", "New complaint received"),
            ("related_complaint_id", "INTEGER", "Foreign Key (complaints), Null", "Related complaint", "1, 2"),
            ("is_read", "BOOLEAN", "Default: False", "Read status", "True, False"),
            ("created_at", "DATETIME", "Auto-populated, Indexed", "Creation timestamp", "2026-02-24 10:00:00"),
            ("action_url", "VARCHAR(255)", "Null", "URL for action", "[url_path]"),
        ],
        "case_statistics": [
            ("id", "INTEGER", "Primary Key", "Unique stat identifier", "1, 2, 3"),
            ("month", "DATE", "Not Null, Indexed", "Stats month", "2026-02-01"),
            ("total_complaints", "INTEGER", "Default: 0", "Total complaints count", "150"),
            ("resolved_complaints", "INTEGER", "Default: 0", "Resolved complaints count", "120"),
            ("pending_complaints", "INTEGER", "Default: 0", "Pending complaints count", "30"),
            ("high_priority_count", "INTEGER", "Default: 0", "High priority count", "25"),
            ("average_resolution_time", "FLOAT", "Null", "Avg resolution time (hours)", "48.5"),
            ("last_updated", "DATETIME", "Auto-updated", "Last update timestamp", "2026-02-24 23:59:59"),
        ],
    }
    
    # Create sheets for each table
    for table_name, fields in tables_definition.items():
        ws = wb.create_sheet(table_name)
        
        # Column headers
        headers = ["Field Name", "Data Type (with Size)", "Constraint", "Description", "Sample Data"]
        for col_num, header_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Data rows
        for row_num, field_data in enumerate(fields, 2):
            for col_num, value in enumerate(field_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 30
        
        # Row height
        ws.row_dimensions[1].height = 35
    
    # Save workbook
    output_file = "d:\\secure-case-connect-main\\DATA_DICTIONARY_UPDATED.xlsx"
    wb.save(output_file)
    
    print("✅ Professional Data Dictionary Created: DATA_DICTIONARY_UPDATED.xlsx")
    print("\n📊 Sheets Created:")
    print("   1. Overview")
    for idx, table_name in enumerate(tables_definition.keys(), 2):
        field_count = len(tables_definition[table_name])
        print(f"   {idx}. {table_name} ({field_count} fields)")
    
    print("\n📈 Each table includes:")
    print("   • Field Name")
    print("   • Data Type (with Size)")
    print("   • Constraint")
    print("   • Description")
    print("   • Sample Data")
    
    print(f"\n📂 Location: {output_file}")
    
    import datetime
    print(f"📅 Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    create_professional_dictionary()
